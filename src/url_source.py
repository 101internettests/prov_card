import os
from typing import List

from openpyxl import load_workbook


def _append_if_valid(urls: List[str], value: str) -> None:
    s = (value or "").strip()
    if s and not s.startswith("#"):
        urls.append(s)


def _load_urls_from_txt(file_path: str) -> List[str]:
    urls: List[str] = []
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            _append_if_valid(urls, line)
    return urls


def _load_urls_from_xlsx(file_path: str) -> List[str]:
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    urls: List[str] = []
    # Собираем первую колонку со всех листов в один список
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            cell = row[0]
            if not cell:
                continue
            _append_if_valid(urls, str(cell))
    return urls


def load_urls(path: str) -> List[str]:
    """Загружает единый список URL из файла (.txt или .xlsx).

    - .txt: по одной ссылке на строку; пустые/комментарии пропускаются
    - .xlsx: первая колонка всех листов; пустые/комментарии пропускаются
    """
    if os.path.isfile(path):
        lower = path.lower()
        if lower.endswith(".txt"):
            return _load_urls_from_txt(path)
        if lower.endswith(".xlsx"):
            return _load_urls_from_xlsx(path)
        raise ValueError("Поддерживаются только .txt или .xlsx файлы для списка URL")

    raise FileNotFoundError(f"Файл со списком URL не найден: {path}")

